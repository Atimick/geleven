import discord
from discord.ext import commands
from discord import app_commands, Interaction, File, Embed
import openpyxl
import os

from .id_check import is_admin, is_mod

class ModCommands(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    def log_command_usage(self, user, command_name, channel_id):
        print(f"[LOG] {user} ({user.id}) dùng lệnh {command_name} ở kênh {channel_id}")

    @app_commands.command(name="mod", description="Moderator Commands")
    @app_commands.check(is_mod)
    async def mod_help(self, interaction: Interaction):
        embed = discord.Embed(
            title="🛡️ Moderator Commands",
            description="Danh sách lệnh dành cho Moderator (bao gồm cả Admin):",
            color=discord.Color.blue()
        )
        embed.add_field(
            name="Quản lý & Xuất dữ liệu",
            value="`/exportchannels`, `/exportmembers`, `/boosters`",
            inline=False
        )
        embed.add_field(
            name="Tiện ích khác",
            value="``/mod`",
            inline=False
        )

        embed.set_footer(text="Geleven Bot • Moderator Panel")
        embed.timestamp = interaction.created_at

        await interaction.response.send_message(embed=embed, ephemeral=True)

    @app_commands.command(name="exportchannels", description="Xuất danh sách kênh và thread trong server")
    @app_commands.check(is_mod)
    async def exportchannels(self, interaction: Interaction):
        self.log_command_usage(interaction.user, "exportchannels", interaction.channel_id)

        if not interaction.guild.me.guild_permissions.view_channel:
            await interaction.response.send_message("❌ Bot cần quyền **Xem kênh (View Channels)** để thực hiện lệnh này!", ephemeral=True)
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Channels"
            ws.append(["Loại chính", "Level", "Tên", "ID", "Tên cha", "ID cha"])

            # Duyệt category
            for category in interaction.guild.categories:
                ws.append(["Category", 0, category.name, str(category.id), "None", "None"])
                for channel in category.channels:
                    if isinstance(channel, discord.TextChannel):
                        ws.append(["Text", 1, channel.name, str(channel.id), category.name, str(category.id)])
                        for thread in channel.threads:
                            ws.append(["Thread", 2, thread.name, str(thread.id), channel.name, str(channel.id)])
                    elif isinstance(channel, discord.VoiceChannel):
                        ws.append(["Voice", 1, channel.name, str(channel.id), category.name, str(category.id)])
                    elif isinstance(channel, discord.ForumChannel):
                        ws.append(["Forum", 1, channel.name, str(channel.id), category.name, str(category.id)])
                        for post in channel.threads:
                            ws.append(["Post", 2, post.name, str(post.id), channel.name, str(channel.id)])

            # Kênh không category
            for channel in interaction.guild.channels:
                if channel.category is None:
                    if isinstance(channel, discord.TextChannel):
                        ws.append(["Text", 1, channel.name, str(channel.id), "None", "None"])
                        for thread in channel.threads:
                            ws.append(["Thread", 2, thread.name, str(thread.id), channel.name, str(channel.id)])
                    elif isinstance(channel, discord.VoiceChannel):
                        ws.append(["Voice", 1, channel.name, str(channel.id), "None", "None"])
                    elif isinstance(channel, discord.ForumChannel):
                        ws.append(["Forum", 1, channel.name, str(channel.id), "None", "None"])
                        for post in channel.threads:
                            ws.append(["Post", 2, post.name, str(post.id), channel.name, str(channel.id)])

            os.makedirs("exports", exist_ok=True)
            file_path = os.path.join("exports", f"channels_{interaction.guild.id}.xlsx")
            wb.save(file_path)

            embed = discord.Embed(
                title="📦 Xuất danh sách kênh hoàn tất",
                description=f"Máy chủ: **{interaction.guild.name}**\n"
                            f"Tổng dòng dữ liệu: **{ws.max_row - 1}**",
                color=discord.Color.blue()
            )
            embed.set_footer(text="Dữ liệu được tạo bởi Geleven Bot")
            embed.timestamp = interaction.created_at

            await interaction.response.send_message(embed=embed, file=discord.File(file_path), ephemeral=True)
            os.remove(file_path)

        except Exception as e:
            await interaction.response.send_message(f"❌ Xuất thất bại: `{e}`", ephemeral=True)

    @app_commands.command(name="exportmembers", description="Xuất danh sách thành viên server")
    @app_commands.check(is_mod)
    async def exportmembers(self, interaction: Interaction):
        self.log_command_usage(interaction.user, "exportmembers", interaction.channel_id)

        bot_member = interaction.guild.me
        if not bot_member.guild_permissions.view_guild_insights and not bot_member.guild_permissions.view_audit_log:
            await interaction.response.send_message("❌ Bot cần quyền `Xem thành viên` hoặc tương đương để xuất danh sách.", ephemeral=True)
            return

        try:
            members = [member async for member in interaction.guild.fetch_members(limit=None)]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Members"
            ws.append(["Display Name", "Username", "User ID", "Created At", "Joined At"])

            for member in members:
                if not member.bot:
                    ws.append([
                        member.display_name,
                        member.name,
                        str(member.id),
                        member.created_at.strftime("%Y-%m-%d"),
                        member.joined_at.strftime("%Y-%m-%d") if member.joined_at else "N/A"
                    ])

            os.makedirs("exports", exist_ok=True)
            file_path = os.path.join("exports", f"members_{interaction.guild.id}.xlsx")
            wb.save(file_path)

            embed = Embed(
                title="📦 Danh sách thành viên đã xuất",
                description=f"Máy chủ: **{interaction.guild.name}**\n"
                            f"Tổng thành viên (không tính bot): **{len([m for m in members if not m.bot])}**",
                color=discord.Color.green()
            )
            embed.set_footer(text="Dữ liệu được tạo tự động bởi Geleven Bot")
            embed.timestamp = interaction.created_at

            await interaction.response.send_message(embed=embed, file=File(file_path), ephemeral=True)
            os.remove(file_path)

        except discord.Forbidden:
            await interaction.response.send_message("❌ Bot không có đủ quyền để xem danh sách thành viên.", ephemeral=True)
        except Exception as e:
            await interaction.response.send_message(f"❌ Xuất thất bại: `{e}`", ephemeral=True)

async def setup(bot: commands.Bot):
    await bot.add_cog(ModCommands(bot))
