import discord
from discord.ext import commands
from discord import app_commands
import random
import re
import textwrap
import secrets
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

LOG_FOLDER = "Geleven/LOG"
LOG_FILE = os.path.join(LOG_FOLDER, "command_log.xlsx")

class PubRoll(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    # --- Slash command ---
    @app_commands.command(name="roll", description="Roll a dice (VD: 1d20, 2d6+3, 1d1000-5)")
    async def roll_slash(self, interaction: discord.Interaction, dice: str):
        await self._handle_roll(dice, interaction.user, interaction=interaction)

    # --- Prefix command ---
    @commands.command(name="r")
    async def roll_prefix(self, ctx, *, dice: str):
        await self._handle_roll(dice, ctx.author, ctx=ctx)

    # --- Xử lý chung (TĂNG WIDTH CỘT SUM VÀ AVG ĐỂ TRÁNH LỒI RA) ---
    async def _handle_roll(self, dice: str, user, interaction: discord.Interaction = None, ctx: commands.Context = None):
        original_dice = dice  # Giữ format gốc user nhập
        is_simple_dice_roll = False
        
        if interaction:
            try:
                await interaction.response.defer()
            except Exception:
                pass

        try:
            dice = dice.strip().lower()
            parts = re.split(r'(\+|-)', dice)

            total = 0
            details = []
            operator = '+'
            
            rolls = []
            sides = 0
            num = 1  # Default

            for part in parts:
                part = part.strip()
                if not part:
                    continue

                if part in ['+', '-']:
                    operator = part
                else:
                    match = re.match(r'(\d*)d(\d+)', part)
                    if match:
                        num = int(match.group(1)) if match.group(1) else 1
                        sides = int(match.group(2))
                        if num <= 0 or sides <= 0:
                            raise ValueError("Số lượng hoặc mặt của xúc xắc không hợp lệ.")

                        current_rolls = [secrets.randbelow(sides) + 1 for _ in range(num)]
                        rolls.extend(current_rolls)
                        
                        subtotal = sum(current_rolls)
                        if operator == '-':
                            subtotal = -subtotal
                        total += subtotal
                        details.append(f"{'-' if operator == '-' else ''}{num}d{sides} ({', '.join(map(str, current_rolls))})")
                    else:
                        try:
                            number = int(part)
                            if operator == '-':
                                number = -number
                            total += number
                            details.append(f"{operator} {abs(number)}")
                        except ValueError:
                            raise ValueError(f"Không hiểu cú pháp: {part}")

            if not details:
                raise ValueError("Không có cú pháp hợp lệ.")
            
            embed = discord.Embed(
                color=discord.Color.dark_purple()
            )
            
            embed.set_author(
                name="Dice Roller", 
                icon_url="https://i.imgur.com/n6t1V2I.png"  # Placeholder
            )
            
            embed.set_footer(text=f"Người roll: {user.display_name}")
            
            # Check cho simple roll
            if len(parts) == 1 and re.match(r'(\d*)d(\d+)', parts[0].strip()):
                roll_result_str = ", ".join(map(str, rolls)) if rolls else " "
                roll_sum_str = str(total)
                average = total / num if num > 0 else 0
                avg_str = f"{average:.2f}"
                
                # Fixed widths (tăng sum và avg lên 8 để đủ cho số lớn)
                ROLLS_COL_WIDTH = 16  # Tăng nhẹ để rolls thoải mái hơn
                SUM_COL_WIDTH = 8     # Tăng từ 6 lên 8
                AVG_COL_WIDTH = 8     # Tăng từ 6 lên 8
                TOTAL_INNER_WIDTH = ROLLS_COL_WIDTH + SUM_COL_WIDTH + AVG_COL_WIDTH + 2  # +2 for two dividers
                
                # Nếu str ngắn, dùng single line
                if len(roll_result_str) <= ROLLS_COL_WIDTH - 2:
                    is_simple_dice_roll = True
                    
                    dice_title_content = original_dice.upper().center(TOTAL_INNER_WIDTH)
                    rolls_title = "rolls".center(ROLLS_COL_WIDTH)
                    sum_title = "sum".center(SUM_COL_WIDTH)
                    avg_title = "avg".center(AVG_COL_WIDTH)
                    roll_content = roll_result_str.center(ROLLS_COL_WIDTH)
                    sum_content_bracketed = f"[{roll_sum_str}]".center(SUM_COL_WIDTH)
                    avg_content_bracketed = f"[{avg_str}]".center(AVG_COL_WIDTH)
                    
                    LINE_FULL = "═" * TOTAL_INNER_WIDTH
                    LINE_ROLLS = "═" * ROLLS_COL_WIDTH
                    LINE_SUM = "═" * SUM_COL_WIDTH
                    LINE_AVG = "═" * AVG_COL_WIDTH
                    
                    table_content = textwrap.dedent(f"""
                        ╔{LINE_FULL}╗
                        ║{dice_title_content}║
                        ╠{LINE_ROLLS}╦{LINE_SUM}╦{LINE_AVG}╣
                        ║{rolls_title}║{sum_title}║{avg_title}║
                        ╠{LINE_ROLLS}╬{LINE_SUM}╬{LINE_AVG}╣
                        ║{roll_content}║{sum_content_bracketed}║{avg_content_bracketed}║
                        ╚{LINE_ROLLS}╩{LINE_SUM}╩{LINE_AVG}╝
                    """)
                    
                    embed.description = f"```\n{table_content}\n```"
                
                else:
                    # Multi-line cho rolls dài
                    roll_lines = textwrap.wrap(roll_result_str, width=ROLLS_COL_WIDTH - 2)
                    max_lines = len(roll_lines)
                    
                    if max_lines > 10:  # Fallback nếu quá dài vertical
                        is_simple_dice_roll = False
                    else:
                        is_simple_dice_roll = True
                        
                        dice_title_content = original_dice.upper().center(TOTAL_INNER_WIDTH)
                        rolls_title = "rolls".center(ROLLS_COL_WIDTH)
                        sum_title = "sum".center(SUM_COL_WIDTH)
                        avg_title = "avg".center(AVG_COL_WIDTH)
                        
                        LINE_FULL = "═" * TOTAL_INNER_WIDTH
                        LINE_ROLLS = "═" * ROLLS_COL_WIDTH
                        LINE_SUM = "═" * SUM_COL_WIDTH
                        LINE_AVG = "═" * AVG_COL_WIDTH
                        
                        # Build content lines
                        content_lines = []
                        sum_content_bracketed = f"[{roll_sum_str}]".center(SUM_COL_WIDTH)
                        avg_content_bracketed = f"[{avg_str}]".center(AVG_COL_WIDTH)
                        for i in range(max_lines):
                            roll_line = roll_lines[i].center(ROLLS_COL_WIDTH)
                            if i == max_lines - 1:
                                sum_line = sum_content_bracketed
                                avg_line = avg_content_bracketed
                            else:
                                sum_line = " " * SUM_COL_WIDTH
                                avg_line = " " * AVG_COL_WIDTH
                            content_lines.append(f"║{roll_line}║{sum_line}║{avg_line}║")
                        
                        table_content = textwrap.dedent(f"""
                            ╔{LINE_FULL}╗
                            ║{dice_title_content}║
                            ╠{LINE_ROLLS}╦{LINE_SUM}╦{LINE_AVG}╣
                            ║{rolls_title}║{sum_title}║{avg_title}║
                            ╠{LINE_ROLLS}╬{LINE_SUM}╬{LINE_AVG}╣
                        """) + "\n".join(content_lines) + f"\n╚{LINE_ROLLS}╩{LINE_SUM}╩{LINE_AVG}╝"
                        
                        embed.description = f"```\n{table_content}\n```"

            if not is_simple_dice_roll:
                # Format text cho complex hoặc rolls quá dài
                result_str = f"**Kết quả:** {total}\n\n**Chi tiết:**\n" + "\n".join(details)
                embed_description = textwrap.shorten(result_str, width=2000, placeholder="...")
                embed.description = embed_description
                embed.add_field(name="Lệnh gốc", value=original_dice, inline=False)

            # Gửi kết quả
            if interaction:
                await interaction.edit_original_response(embed=embed)
                self.log_command_usage(user, "roll", interaction.channel.id)
            elif ctx:
                await ctx.send(embed=embed)
                self.log_command_usage(user, "roll", ctx.channel.id)

        except Exception as e:
            error_msg = f"❌ Lỗi: {str(e)}"
            if interaction:
                await interaction.edit_original_response(content=error_msg, embed=None)
            elif ctx:
                await ctx.send(error_msg)

    # --- Ghi log ---
    def log_command_usage(self, user, command_name, channel_id):
        os.makedirs(LOG_FOLDER, exist_ok=True)

        if not os.path.exists(LOG_FILE):
            wb = Workbook()
            ws = wb.active
            ws.append(["Timestamp", "User", "UserID", "Command", "ChannelID"])
            wb.save(LOG_FILE)

        wb = load_workbook(LOG_FILE)
        ws = wb.active
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            str(user),
            str(user.id),
            command_name,
            str(channel_id)
        ])
        wb.save(LOG_FILE)

async def setup(bot: commands.Bot):
    await bot.add_cog(PubRoll(bot))