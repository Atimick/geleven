import discord
import openpyxl
import os
from datetime import datetime, timedelta, timezone
from discord import app_commands, Interaction, Embed, File, ChannelType
from discord.ext import commands

from .id_check import is_admin, is_mod, is_allowed, LOG_FOLDER, BOOST_LOG_PATH

class AMCC(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    def log_command_usage(self, user, command_name, channel_id):
        os.makedirs(LOG_FOLDER, exist_ok=True)
        file_path = os.path.join(LOG_FOLDER, "command_log.xlsx")

        if not os.path.exists(file_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Command Usage"
            ws.append(["User", "User ID", "Command", "Channel ID", "Timestamp"])
            wb.save(file_path)

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        ws.append([
            user.display_name,
            str(user.id),
            command_name,
            str(channel_id),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])
        wb.save(file_path)

    @app_commands.command(name="adm", description="Hiện danh sách lệnh Quản trị")
    @app_commands.check(is_allowed) 
    async def adm(self, interaction: Interaction):
        self.log_command_usage(interaction.user, "adm", interaction.channel_id)

        embed = discord.Embed(
            title="🛠️ Danh sách lệnh Quản trị (Admin/Mod)",
            description="Các lệnh này chỉ dành cho người có quyền hạn phù hợp.",
            color=discord.Color.orange()
        )

        embed.add_field(
            name="🔧 Lệnh được Ati cấp quyền sử dụng:",
            value=(
                "• `/adm` — (Mod/Admin) Hiện danh sách các lệnh admin\n"
                "• `/boosters` — 💎 (Admin) Liệt kê thành viên đang Boost server\n"
                "• `/mdk` — (Admin) Thanh lý hàng (Us)\n"
                "• `/mdc` — (Admin) Thanh lý kho (Cs)\n"
                "• `/purgechat` — 📄 (Admin) Xóa tin nhắn hàng loạt"
            ),
            inline=False
        )
        if is_admin(interaction):
            embed.add_field(
                name="🛡️ Lệnh đặc biệt:",
                value=(
                    "• `/panpakapan` — 👑 Gán role 'Panpakapan' với quyền admin\n"
                    "• `/nopan` — ❌ Xóa role 'Panpakapan'\n"
                    "• **`/reload`** — 🔄 Tải lại các module (Cog) của bot **(Chỉ Ati Admin)**"
                ),
                inline=False
            )

        embed.set_footer(text=f"Yêu cầu bởi {interaction.user.display_name}", icon_url=interaction.user.display_avatar.url)
        await interaction.response.send_message(embed=embed, ephemeral=True)

    @app_commands.command(name="panpakapan", description="Tạo và gán role 'Panpakapan' với quyền admin (chỉ Admin)")
    @app_commands.check(is_admin)
    async def panpakapan(self, interaction: Interaction):
        if interaction.guild and interaction.guild.id == 733363418681049159:
            await interaction.response.send_message("\u26a0\ufe0f Server AM denied.", ephemeral=True)
            return

        self.log_command_usage(interaction.user, "panpakapan", interaction.channel_id)

        guild = interaction.guild
        if not guild.me.guild_permissions.administrator:
            await interaction.response.send_message("\u274c Bot cần quyền Administrator để tạo role!", ephemeral=True)
            return

        existing_role = discord.utils.get(guild.roles, name="Panpakapan")
        if existing_role:
            await interaction.response.send_message("\ud83d\udccc Role 'Panpakapan' exists.", ephemeral=True)
        else:
            permissions = discord.Permissions(administrator=True)
            new_role = await guild.create_role(name="Panpakapan", permissions=permissions)
            await interaction.user.add_roles(new_role)
            await interaction.response.send_message("\u2705 Role 'Panpakapan' added.", ephemeral=True)

    @app_commands.command(name="nopan", description="Xóa role 'Panpakapan' (chỉ Admin)")
    @app_commands.check(is_admin)
    async def nopan(self, interaction: Interaction):
        self.log_command_usage(interaction.user, "nopan", interaction.channel_id)

        guild = interaction.guild
        if not guild.me.guild_permissions.administrator:
            await interaction.response.send_message("\u274c Bot cần quyền Administrator để xóa role!", ephemeral=True)
            return

        role = discord.utils.get(guild.roles, name="Panpakapan")
        if role:
            try:
                await role.delete()
                await interaction.response.send_message("\ud83d\uddd1\ufe0f Deleted 'Panpakapan'.", ephemeral=True)
            except discord.Forbidden:
                await interaction.response.send_message("\u274c Bot không có quyền xóa role!", ephemeral=True)
            except Exception as e:
                await interaction.response.send_message(f"\u274c Lỗi: {e}", ephemeral=True)
        else:
            await interaction.response.send_message("\u26a0\ufe0f Không tìm thấy 'Panpakapan'.", ephemeral=True)

    @app_commands.command(name="boosters", description="Hiển thị danh sách boosters và xuất ra file excel")
    @app_commands.check(is_mod)
    async def boosters(self, interaction: Interaction):
        self.log_command_usage(interaction.user, "boosters", interaction.channel_id)
        
        bot_member = interaction.guild.me
        if not bot_member.guild_permissions.view_guild_insights and not bot_member.guild_permissions.view_audit_log:
            await interaction.response.send_message("❌ Bot cần quyền **Xem nhật ký kiểm duyệt** (View Audit Log) để thực hiện lệnh này.", ephemeral=True)
            return
        
        await interaction.response.defer(thinking=True, ephemeral=True)

        boosters = []
        async for member in interaction.guild.fetch_members(limit=None):
            if member.premium_since:
                boost_utc = member.premium_since
                now_utc = datetime.utcnow().replace(tzinfo=timezone.utc)
                boost_vn = boost_utc + timedelta(hours=7)
                days = (now_utc - boost_utc).days
                boosters.append((member, days, boost_utc, boost_vn))
        
        if not boosters:
            await interaction.followup.send("⚠️ Không có ai đang Boost server.", ephemeral=True)
            return

        embed = Embed(
            title="💎 Danh sách Boosters",
            description=f"Hiện có `{len(boosters)}` thành viên đang Boost server.",
            color=discord.Color.purple()
        )
        embed.set_footer(text=f"Export bởi: {interaction.user.display_name}", icon_url=interaction.user.display_avatar.url)

        for member, days, boost_utc, boost_vn in boosters[:10]:
            embed.add_field(
                name=f"{member.display_name} ({days} ngày)",
                value=f"• Boost từ (UTC): `{boost_utc.strftime('%d/%m/%Y %H:%M:%S')}`\n• Boost từ (VN): `{boost_vn.strftime('%d/%m/%Y %H:%M:%S')}`",
                inline=False
            )

        if len(boosters) > 10:
            embed.add_field(name="...", value=f"Và {len(boosters) - 10} người khác nữa.", inline=False)
            
        await interaction.followup.send(embed=embed, ephemeral=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Boosters"
        ws.append(["Display Name", "Username", "User ID", "Boost Start (UTC)", "Boost Start (VN)", "Boost Days"])

        for m, d, utc, vn in boosters:
            ws.append([m.display_name, m.name, str(m.id), utc.strftime('%d/%m/%Y %H:%M:%S'), vn.strftime('%d/%m/%Y %H:%M:%S'), d])

        os.makedirs(os.path.dirname(BOOST_LOG_PATH), exist_ok=True)
        wb.save(BOOST_LOG_PATH)

        await interaction.followup.send(content="📄 File danh sách Boosters:", file=File(BOOST_LOG_PATH), ephemeral=True)

    @app_commands.command(name="mdk", description="Thanh lý hàng")
    @app_commands.check(is_admin)
    async def masskick(self, interaction: Interaction):
        self.log_command_usage(interaction.user, "mdk", interaction.channel_id)

        if not interaction.guild.me.guild_permissions.kick_members:
            await interaction.response.send_message("\u274c Bot cần quyền kick thành viên!", ephemeral=True)
            return

        file_path = os.path.join(os.path.dirname(__file__), "..", "ID-Kicklist.txt")

        if not os.path.exists(file_path):
            await interaction.response.send_message(f"❌ Không tìm thấy file: `{file_path}`", ephemeral=True)
            return

        try:
            with open(file_path, "r", encoding="utf-8") as f:
                user_ids = [int(line.strip()) for line in f if line.strip().isdigit()]
        except Exception as e:
            await interaction.response.send_message(f"❌ Không đọc được file: {e}", ephemeral=True)
            return

        if not user_ids:
            await interaction.response.send_message("❌ File không chứa ID hợp lệ để thanh lý!", ephemeral=True)
            return

        kicked = []
        failed = []

        for uid in user_ids:
            member = interaction.guild.get_member(uid)
            if member:
                try:
                    await member.kick(reason="Thanh lý hàng (mass kick by admin)")
                    kicked.append(f"{member.display_name} ({member.id})")
                except Exception as e:
                    failed.append((f"{member.display_name} ({member.id})", str(e)))
            else:
                failed.append((str(uid), "Không tìm thấy trong server"))

        msg = ""
        if kicked:
            msg += f"✅ Đã thanh lý {len(kicked)} hàng:\n" + "\n".join(kicked)
        if failed:
            msg += f"\n\n❌ Lỗi khi xử lý {len(failed)} hàng:\n" + "\n".join(f"{name} ({reason})" for name, reason in failed)

        await interaction.response.send_message(f"**Kết quả Thanh lý hàng:**\n```{msg}```", ephemeral=True)

    @app_commands.command(name="mdc", description="Thanh lý kho")
    @app_commands.check(is_admin)
    async def masschanneldelete(self, interaction: Interaction):
        self.log_command_usage(interaction.user, "mdc", interaction.channel_id)

        if not interaction.guild.me.guild_permissions.manage_channels:
            await interaction.response.send_message("❌ Bot cần quyền **quản lý kênh**!", ephemeral=True)
            return

        file_path = os.path.join(os.path.dirname(__file__), "..", "ID-ChannelDelete.txt")

        if not os.path.exists(file_path):
            await interaction.response.send_message(f"❌ Không tìm thấy file: `{file_path}`", ephemeral=True)
            return

        try:
            with open(file_path, "r", encoding="utf-8") as f:
                channel_ids = [int(line.strip()) for line in f if line.strip().isdigit()]
        except Exception as e:
            await interaction.response.send_message(f"❌ Không thể đọc file: {e}", ephemeral=True)
            return

        if not channel_ids:
            await interaction.response.send_message("❌ File không chứa ID hợp lệ để xóa!", ephemeral=True)
            return

        deleted = []
        failed = []

        for cid in channel_ids:
            channel = interaction.guild.get_channel(cid)
            if channel:
                try:
                    await channel.delete(reason="Thanh lý kho (mass delete by admin)")
                    deleted.append(f"{channel.name} ({channel.id})")
                except Exception as e:
                    failed.append((f"{channel.name} ({channel.id})", str(e)))
            else:
                failed.append((str(cid), "Không tìm thấy trong server"))

        msg = ""
        if deleted:
            msg += f"✅ Đã thanh lý {len(deleted)} kho:\n" + "\n".join(deleted)
        if failed:
            msg += f"\n\n❌ Lỗi khi xử lý {len(failed)} kho:\n" + "\n".join(f"{name} ({reason})" for name, reason in failed)

        await interaction.response.send_message(f"**Kết quả Thanh lý kho:**\n```{msg}```", ephemeral=True)
    
    @app_commands.command(name="purgechat", description="Xóa tin nhắn trong kênh chat, bao gồm cả kênh voice chat.")
    @app_commands.check(is_admin)
    async def purgechat(self, interaction: Interaction):
        """
        Xóa tất cả tin nhắn trong kênh hiện tại ngược về quá khứ 14 ngày,
        không xóa tin nhắn đã ghim và tin nhắn gửi sau khi lệnh được gọi.
        """
        self.log_command_usage(interaction.user, "purgechat", interaction.channel_id)

        # Lấy kênh để xử lý. Kiểm tra cả kênh văn bản thông thường và kênh văn bản của voice chat.
        target_channel = interaction.channel
        if not isinstance(target_channel, (discord.TextChannel, discord.VoiceChannel)):
            await interaction.response.send_message(
                "❌ Lệnh này chỉ có thể được sử dụng trong các kênh chat hoặc kênh voice có chat.",
                ephemeral=True
            )
            return

        # Nếu lệnh được gọi trong kênh voice, sử dụng kênh văn bản đi kèm của nó
        if isinstance(target_channel, discord.VoiceChannel):
            if not target_channel.type == ChannelType.voice:
                # Đây là kênh voice nhưng không có text chat
                # (trường hợp hiếm, nhưng nên xử lý)
                await interaction.response.send_message(
                    "❌ Kênh voice này không hỗ trợ chat văn bản.",
                    ephemeral=True
                )
                return
            pass

        await interaction.response.defer(thinking=True, ephemeral=True)

        try:
            # Lấy thời điểm lệnh được gọi làm mốc
            deleted_messages = await target_channel.purge(
                limit=None,
                before=interaction.created_at,
                check=lambda m: not m.pinned
            )
            
            await interaction.followup.send(
                f"🧹 Đã dọn dẹp thành công **{len(deleted_messages)}** tin nhắn.",
                ephemeral=True
            )

        except discord.Forbidden:
            await interaction.followup.send(
                "❌ Bot không có quyền **Quản lý tin nhắn** trong kênh này.",
                ephemeral=True
            )
        except Exception as e:
            await interaction.followup.send(
                f"❌ Đã xảy ra lỗi không mong muốn: `{e}`",
                ephemeral=True
            )

async def setup(bot):
    await bot.add_cog(AMCC(bot))